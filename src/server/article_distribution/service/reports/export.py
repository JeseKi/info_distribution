# -*- coding: utf-8 -*-
"""Overview report export services."""

from __future__ import annotations

import csv
from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Literal, Sequence

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ...schemas import (
    AccountStatusFilter,
    ArticleDistributionOverviewArticleOut,
    ArticleDistributionOverviewTopicOut,
    ArticleDistributionOverviewUserOut,
    ArticleDistributionOverviewView,
)
from .overview import list_report_overview

OverviewExportFormat = Literal["csv", "xlsx"]

EXPORT_MEDIA_TYPES = {
    "csv": "text/csv; charset=utf-8-sig",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

EXPORT_FILE_EXTENSIONS = {
    "csv": "csv",
    "xlsx": "xlsx",
}

_MAX_EXPORT_ROWS = 1000000

_PUBLICATION_TYPE_LABELS = {
    "video": "视频",
    "article": "文章",
    "image_text": "图文",
}

_PUBLISH_STATUS_LABELS = {
    "unpublished": "未发布",
    "published": "已发布",
    "invalid": "文档失效",
}

_VIEW_SHEET_NAMES = {
    "users": "用户汇总",
    "articles": "文章明细",
    "topics": "选题汇总",
}


def build_report_overview_export(
    db: Session,
    *,
    export_format: OverviewExportFormat,
    view: ArticleDistributionOverviewView = "users",
    keyword: str | None = None,
    scheduled_from: date | None = None,
    scheduled_to: date | None = None,
    platform: str | None = None,
    publication_type: str | None = None,
    account_status: AccountStatusFilter = "active",
    publish_status: str | None = None,
    missing_traffic_only: bool = False,
    recorded_from: datetime | None = None,
    recorded_to: datetime | None = None,
) -> bytes:
    report = list_report_overview(
        db,
        view=view,
        keyword=keyword,
        scheduled_from=scheduled_from,
        scheduled_to=scheduled_to,
        platform=platform,
        publication_type=publication_type,
        account_status=account_status,
        publish_status=publish_status,
        missing_traffic_only=missing_traffic_only,
        recorded_from=recorded_from,
        recorded_to=recorded_to,
        page=1,
        page_size=_MAX_EXPORT_ROWS,
    )
    rows = _rows_for_view(view, report.items)
    if export_format == "csv":
        return _build_csv(rows)
    return _build_xlsx(_VIEW_SHEET_NAMES[view], rows)


def _rows_for_view(view: str, items: Sequence[object]) -> list[list[object]]:
    if view == "articles":
        return _article_rows(
            [
                item
                for item in items
                if isinstance(item, ArticleDistributionOverviewArticleOut)
            ]
        )
    if view == "topics":
        return _topic_rows(
            [
                item
                for item in items
                if isinstance(item, ArticleDistributionOverviewTopicOut)
            ]
        )
    return _user_rows(
        [
            item
            for item in items
            if isinstance(item, ArticleDistributionOverviewUserOut)
        ]
    )


def _user_rows(users: list[ArticleDistributionOverviewUserOut]) -> list[list[object]]:
    rows: list[list[object]] = [
        [
            "用户ID",
            "负责人",
            "用户名",
            "邮箱",
            "剩余未发布",
            "已发布",
            "失效",
            "停用账号文章",
            "未填流量",
            "阅读量",
            "点赞量",
            "收藏量",
            "转发量",
        ]
    ]
    for user in users:
        rows.append(
            [
                user.user_id,
                user.name or user.username,
                user.username,
                user.email,
                user.remaining_count,
                user.published_count,
                user.invalid_count,
                user.inactive_account_articles,
                user.missing_count,
                user.read_count,
                user.like_count,
                user.favorite_count,
                user.share_count,
            ]
        )
    return rows


def _article_rows(
    articles: list[ArticleDistributionOverviewArticleOut],
) -> list[list[object]]:
    rows: list[list[object]] = [
        [
            "文章ID",
            "标题",
            "负责人ID",
            "负责人",
            "用户名",
            "邮箱",
            "平台",
            "发布账号",
            "发布类型",
            "账号状态",
            "发布状态",
            "计划日期",
            "未填流量",
            "选题",
            "Output ID",
            "角色",
            "角度",
            "受众",
            "发布链接",
            "阅读量",
            "点赞量",
            "收藏量",
            "转发量",
            "统计时间",
        ]
    ]
    for article in articles:
        latest_stat = article.latest_traffic_stat
        rows.append(
            [
                article.id,
                article.title,
                article.user_id,
                article.name or article.username,
                article.username,
                article.email,
                article.platform,
                article.account_name,
                _PUBLICATION_TYPE_LABELS.get(
                    article.publication_type, article.publication_type
                ),
                "启用" if article.account_is_active else "停用",
                _PUBLISH_STATUS_LABELS.get(article.publish_status, article.publish_status),
                article.scheduled_date.isoformat(),
                "是" if article.missing_traffic else "否",
                article.topic or "",
                article.output_id or "",
                article.article_role or "",
                article.angle_label or "",
                article.audience_label or "",
                article.published_url or "",
                latest_stat.read_count if latest_stat is not None else "",
                latest_stat.like_count if latest_stat is not None else "",
                latest_stat.favorite_count if latest_stat is not None else "",
                latest_stat.share_count if latest_stat is not None else "",
                latest_stat.recorded_at.isoformat() if latest_stat is not None else "",
            ]
        )
    return rows


def _topic_rows(topics: list[ArticleDistributionOverviewTopicOut]) -> list[list[object]]:
    rows: list[list[object]] = [
        [
            "Output ID",
            "选题",
            "素材",
            "文章数",
            "阅读量",
            "点赞量",
            "收藏量",
            "转发量",
        ]
    ]
    for topic in topics:
        rows.append(
            [
                topic.output_id or "",
                topic.topic,
                "\n".join(topic.materials),
                topic.article_count,
                topic.read_count,
                topic.like_count,
                topic.favorite_count,
                topic.share_count,
            ]
        )
    return rows


def _build_csv(rows: list[list[object]]) -> bytes:
    output = StringIO(newline="")
    writer = csv.writer(output)
    writer.writerows(rows)
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def _build_xlsx(sheet_name: str, rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    assert worksheet is not None
    worksheet.title = sheet_name
    for row in rows:
        worksheet.append(row)

    for column_index, column in enumerate(worksheet.columns, start=1):
        max_length = max(len(str(cell.value or "")) for cell in column)
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max_length + 2, 10),
            48,
        )

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
