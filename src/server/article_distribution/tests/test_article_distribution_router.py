# -*- coding: utf-8 -*-
"""Article distribution router tests."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO
import socket

from openpyxl import load_workbook
import pytest
from fastapi import HTTPException
from sqlalchemy.orm import Session

from src.server.article_distribution import router as article_distribution_router
from src.server.article_distribution.models import (
    ArticleDistributionAPIKey,
    ArticleDistributionArticle,
)
from src.server.auth.models import User
from src.server.auth.schemas import UserRole
from src.server.auth import service as auth_service
from src.server.project_management.dao import ProjectManagementDAO
from src.server.project_management.models import Project, Theme
from src.server.project_management.service import bootstrap_default_project_theme


def _create_user(
    db: Session,
    *,
    username: str,
    role: UserRole = UserRole.USER,
    name: str | None = None,
    wechat_nickname: str | None = None,
    wechat_id: str | None = None,
) -> User:
    user = User(
        username=username,
        email=f"{username}@example.com",
        name=username if name is None else name,
        wechat_nickname=wechat_nickname,
        wechat_id=wechat_id,
        role=role,
    )
    user.set_password("Password123")
    db.add(user)
    db.commit()
    db.refresh(user)
    project = bootstrap_default_project_theme(db)
    ProjectManagementDAO(db).add_user_project(user.id, project.id)
    db.refresh(user)
    return user


def _headers(user: User) -> dict[str, str]:
    token = auth_service.create_access_token(
        {
            "sub": user.username,
            "scope": auth_service.get_user_scopes(user),
            "tv": user.token_version,
        }
    )
    return {"Authorization": f"Bearer {token}"}


def test_image_proxy_allows_trusted_fstc_private_address(monkeypatch):
    def fake_getaddrinfo(host: str, port):
        assert host == "fstc.kispace.cc"
        assert port is None
        return [
            (
                socket.AF_INET,
                socket.SOCK_STREAM,
                6,
                "",
                ("172.29.0.33", 0),
            )
        ]

    monkeypatch.setattr(
        article_distribution_router.socket, "getaddrinfo", fake_getaddrinfo
    )

    url = "https://fstc.kispace.cc/i/8b2737602828b9d730105b75fb5f5309.jpg"
    assert article_distribution_router._validate_proxy_image_url(url) == url


def test_image_proxy_rejects_untrusted_private_address(monkeypatch):
    def fake_getaddrinfo(host: str, port):
        assert host == "private.example.com"
        assert port is None
        return [
            (
                socket.AF_INET,
                socket.SOCK_STREAM,
                6,
                "",
                ("172.29.0.33", 0),
            )
        ]

    monkeypatch.setattr(
        article_distribution_router.socket, "getaddrinfo", fake_getaddrinfo
    )

    with pytest.raises(HTTPException) as exc_info:
        article_distribution_router._validate_proxy_image_url(
            "https://private.example.com/image.jpg"
        )

    assert exc_info.value.status_code == 400
    assert exc_info.value.detail == "不允许代理内网或本机图片地址"


def test_v2_account_directory_groups_accounts_by_user_with_api_key(
    test_client, test_db_session: Session
):
    owner_a = _create_user(test_db_session, username="owner_a", name="Owner A")
    owner_b = _create_user(test_db_session, username="owner_b", name="Owner B")
    owner_without_name = _create_user(
        test_db_session, username="owner_no_name", name=""
    )
    admin = _create_user(test_db_session, username="directory_admin", role=UserRole.ADMIN)

    created_accounts = []
    for user_id, publication_type in [
        (owner_a.id, "image_text"),
        (owner_a.id, "article"),
        (owner_b.id, "image_text"),
        (owner_without_name.id, "video"),
    ]:
        create_resp = test_client.post(
            "/api/article-distribution/accounts",
            headers=_headers(admin),
            json={
                "user_id": user_id,
                "account_name": "主号",
                "platform": "wechat",
                "publication_type": publication_type,
            },
        )
        assert create_resp.status_code == 201
        created_accounts.append(create_resp.json())

    key_resp = test_client.post(
        "/api/admin/article-distribution/api-keys",
        headers=_headers(admin),
        json={"name": "directory"},
    )
    assert key_resp.status_code == 201

    directory_resp = test_client.get(
        "/api/v2/article-distribution/accounts",
        headers={"X-API-Key": key_resp.json()["api_key"]},
    )

    assert directory_resp.status_code == 200
    assert directory_resp.json() == [
        {
            "id": owner_a.id,
            "name": "Owner A",
            "accounts": [
                {
                    "id": created_accounts[1]["id"],
                    "project_ids": created_accounts[1]["project_ids"],
                    "projects": created_accounts[1]["projects"],
                    "theme_id": created_accounts[1]["theme_id"],
                    "theme": created_accounts[1]["theme"],
                    "platform": "wechat",
                    "account_name": "主号",
                    "publication_type": "article",
                    "is_active": True,
                },
                {
                    "id": created_accounts[0]["id"],
                    "project_ids": created_accounts[0]["project_ids"],
                    "projects": created_accounts[0]["projects"],
                    "theme_id": created_accounts[0]["theme_id"],
                    "theme": created_accounts[0]["theme"],
                    "platform": "wechat",
                    "account_name": "主号",
                    "publication_type": "image_text",
                    "is_active": True,
                },
            ],
        },
        {
            "id": owner_b.id,
            "name": "Owner B",
            "accounts": [
                {
                    "id": created_accounts[2]["id"],
                    "project_ids": created_accounts[2]["project_ids"],
                    "projects": created_accounts[2]["projects"],
                    "theme_id": created_accounts[2]["theme_id"],
                    "theme": created_accounts[2]["theme"],
                    "platform": "wechat",
                    "account_name": "主号",
                    "publication_type": "image_text",
                    "is_active": True,
                },
            ],
        },
        {
            "id": owner_without_name.id,
            "name": "owner_no_name",
            "accounts": [
                {
                    "id": created_accounts[3]["id"],
                    "project_ids": created_accounts[3]["project_ids"],
                    "projects": created_accounts[3]["projects"],
                    "theme_id": created_accounts[3]["theme_id"],
                    "theme": created_accounts[3]["theme"],
                    "platform": "wechat",
                    "account_name": "主号",
                    "publication_type": "video",
                    "is_active": True,
                },
            ],
        },
    ]

    options_resp = test_client.get(
        "/api/v2/article-distribution/project-themes",
        headers={"X-API-Key": key_resp.json()["api_key"]},
    )
    assert options_resp.status_code == 200
    options = options_resp.json()
    assert options["projects"][0]["id"] == created_accounts[0]["project_ids"][0]
    assert options["projects"][0]["code"] == created_accounts[0]["projects"][0]["code"]
    assert options["themes"][0]["id"] == created_accounts[0]["theme_id"]


def test_v1_endpoints_are_deprecated(test_client):
    account_resp = test_client.get("/api/v1/article-distribution/accounts")
    assert account_resp.status_code == 410

    upload_resp = test_client.post(
        "/api/v1/article-distribution/articles",
        json={
            "account_id": 1,
            "articles": [
                {
                    "title": "Deprecated",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-21",
                    "project_id": 1,
                }
            ],
        },
    )
    assert upload_resp.status_code == 410

    update_resp = test_client.patch(
        "/api/v1/article-distribution/articles/1",
        json={"title": "Deprecated"},
    )
    assert update_resp.status_code == 410


def test_user_can_manage_own_accounts(test_client, test_db_session: Session):
    user = _create_user(test_db_session, username="dist_user")

    create_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(user),
        json={
            "account_name": "主号",
            "platform": "知乎",
            "publication_type": "article",
        },
    )

    assert create_resp.status_code == 201
    account = create_resp.json()
    assert account["user_id"] == user.id
    assert account["account_name"] == "主号"

    list_resp = test_client.get(
        "/api/article-distribution/accounts", headers=_headers(user)
    )
    assert list_resp.status_code == 200
    assert [item["id"] for item in list_resp.json()] == [account["id"]]


def test_user_account_page_is_scoped_to_current_user(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="account_page_owner")
    other = _create_user(test_db_session, username="account_page_other")
    admin = _create_user(test_db_session, username="account_page_admin", role=UserRole.ADMIN)

    owner_accounts = []
    for platform, account_name, publication_type, is_active in [
        ("知乎", "主号", "article", True),
        ("公众号", "备用号", "image_text", False),
    ]:
        create_resp = test_client.post(
            "/api/article-distribution/accounts",
            headers=_headers(owner),
            json={
                "account_name": account_name,
                "platform": platform,
                "publication_type": publication_type,
                "is_active": is_active,
            },
        )
        assert create_resp.status_code == 201
        owner_accounts.append(create_resp.json())

    other_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": other.id,
            "account_name": "主号",
            "platform": "知乎",
            "publication_type": "article",
        },
    )
    assert other_resp.status_code == 201

    page_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(owner),
        params={"page": 1, "page_size": 10},
    )

    assert page_resp.status_code == 200
    page_data = page_resp.json()
    assert page_data["total"] == 2
    assert page_data["page"] == 1
    assert page_data["page_size"] == 10
    assert {item["id"] for item in page_data["items"]} == {
        account["id"] for account in owner_accounts
    }

    keyword_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(owner),
        params={"keyword": "备用", "is_active": False},
    )
    assert keyword_resp.status_code == 200
    assert [item["account_name"] for item in keyword_resp.json()["items"]] == ["备用号"]

    denied_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(owner),
        params={"user_id": other.id},
    )
    assert denied_resp.status_code == 403


def test_admin_account_page_filters_accounts(
    test_client, test_db_session: Session
):
    owner_a = _create_user(test_db_session, username="account_filter_owner_a")
    owner_b = _create_user(test_db_session, username="account_filter_owner_b")
    admin = _create_user(test_db_session, username="account_filter_admin", role=UserRole.ADMIN)

    for user_id, platform, account_name, publication_type, is_active in [
        (owner_a.id, "知乎", "主号", "article", True),
        (owner_a.id, "公众号", "图文号", "image_text", True),
        (owner_b.id, "B站", "视频号", "video", False),
        (owner_b.id, "知乎", "矩阵号", "article", True),
    ]:
        create_resp = test_client.post(
            "/api/article-distribution/accounts",
            headers=_headers(admin),
            json={
                "user_id": user_id,
                "account_name": account_name,
                "platform": platform,
                "publication_type": publication_type,
                "is_active": is_active,
            },
        )
        assert create_resp.status_code == 201

    all_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(admin),
        params={"page": 1, "page_size": 2},
    )
    assert all_resp.status_code == 200
    all_data = all_resp.json()
    assert all_data["total"] == 4
    assert len(all_data["items"]) == 2

    user_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(admin),
        params={"user_id": owner_a.id},
    )
    assert user_resp.status_code == 200
    assert user_resp.json()["total"] == 2
    assert {item["user_id"] for item in user_resp.json()["items"]} == {owner_a.id}

    filtered_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(admin),
        params={
            "platform": "知乎",
            "publication_type": "article",
            "is_active": True,
            "keyword": "矩阵",
        },
    )
    assert filtered_resp.status_code == 200
    filtered_data = filtered_resp.json()
    assert filtered_data["total"] == 1
    assert filtered_data["items"][0]["account_name"] == "矩阵号"
    assert filtered_data["items"][0]["user_id"] == owner_b.id


def test_account_page_filters_by_project_and_theme(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="account_project_filter_owner")
    other = _create_user(test_db_session, username="account_project_filter_other")
    admin = _create_user(
        test_db_session,
        username="account_project_filter_admin",
        role=UserRole.ADMIN,
    )
    dao = ProjectManagementDAO(test_db_session)
    default_project = dao.get_project_by_name("AIFC")
    default_theme = dao.get_theme_by_name("AI")
    assert default_project is not None
    assert default_theme is not None

    second_theme = dao.create_theme(Theme(name="账号筛选主题", is_active=True))
    second_project = dao.create_project(
        Project(name="账号筛选项目", code="ACCTFILT", is_active=True)
    )
    dao.replace_project_themes(second_project.id, [second_theme.id])
    dao.add_user_project(owner.id, second_project.id)
    dao.add_user_project(other.id, second_project.id)

    default_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        json={
            "account_name": "默认项目号",
            "platform": "公众号",
            "publication_type": "article",
            "theme_id": default_theme.id,
        },
    )
    assert default_resp.status_code == 201
    default_account = default_resp.json()

    second_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        json={
            "account_name": "第二项目号",
            "platform": "知乎",
            "publication_type": "article",
            "theme_id": second_theme.id,
        },
    )
    assert second_resp.status_code == 201
    second_account = second_resp.json()

    other_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": other.id,
            "account_name": "他人第二项目号",
            "platform": "知乎",
            "publication_type": "article",
            "theme_id": second_theme.id,
        },
    )
    assert other_resp.status_code == 201
    other_account = other_resp.json()

    project_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(owner),
        params={"project_id": second_project.id},
    )
    assert project_resp.status_code == 200
    project_data = project_resp.json()
    assert project_data["total"] == 1
    assert project_data["items"][0]["id"] == second_account["id"]
    assert project_data["items"][0]["project_ids"] == [second_project.id]

    theme_resp = test_client.get(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        params={"theme_id": default_theme.id},
    )
    assert theme_resp.status_code == 200
    assert [item["id"] for item in theme_resp.json()] == [default_account["id"]]

    empty_intersection_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(owner),
        params={"project_id": default_project.id, "theme_id": second_theme.id},
    )
    assert empty_intersection_resp.status_code == 200
    assert empty_intersection_resp.json()["total"] == 0
    assert empty_intersection_resp.json()["items"] == []

    admin_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(admin),
        params={"project_id": second_project.id, "theme_id": second_theme.id},
    )
    assert admin_resp.status_code == 200
    admin_data = admin_resp.json()
    assert admin_data["total"] == 2
    assert {item["id"] for item in admin_data["items"]} == {
        second_account["id"],
        other_account["id"],
    }

    denied_resp = test_client.get(
        "/api/article-distribution/accounts/page",
        headers=_headers(owner),
        params={"user_id": other.id, "project_id": second_project.id},
    )
    assert denied_resp.status_code == 403


def test_user_cannot_access_other_users_articles(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="owner")
    other = _create_user(test_db_session, username="other")
    admin = _create_user(test_db_session, username="article_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        json={
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "image_text",
        },
    )
    account_id = account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "A",
                    "markdown_content": "# A",
                    "scheduled_date": "2026-05-20",
                    "project_id": 1,
                }
            ],
        },
    )
    assert upload_resp.status_code == 201
    article_id = upload_resp.json()[0]["id"]

    denied_resp = test_client.get(
        f"/api/article-distribution/articles/{article_id}", headers=_headers(other)
    )
    assert denied_resp.status_code == 403


def test_admin_and_api_key_can_upload_articles(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="api_owner")
    admin = _create_user(test_db_session, username="api_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "知乎号",
            "platform": "zhihu",
            "publication_type": "article",
        },
    )
    assert account_resp.status_code == 201
    account = account_resp.json()
    account_id = account["id"]

    key_resp = test_client.post(
        "/api/admin/article-distribution/api-keys",
        headers=_headers(admin),
        json={"name": "integration"},
    )
    assert key_resp.status_code == 201
    raw_key = key_resp.json()["api_key"]
    assert raw_key.startswith("adv1_")

    api_resp = test_client.post(
        "/api/v2/article-distribution/articles",
        headers={"X-API-Key": raw_key},
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "API article",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-21",
                    "theme_id": account["theme_id"],
                },
                {
                    "title": "API article 2",
                    "markdown_content": "body2",
                    "scheduled_date": "2026-05-21",
                    "theme_id": account["theme_id"],
                },
            ],
        },
    )
    assert api_resp.status_code == 201
    data = api_resp.json()
    assert len(data) == 2
    assert data[0]["user_id"] == owner.id
    assert data[0]["project_id"] == account["project_ids"][0]
    assert data[0]["source"] == "api"

    stored_key = test_db_session.query(ArticleDistributionAPIKey).first()
    assert stored_key is not None
    assert stored_key.last_used_at is not None


def test_v2_api_key_can_update_article_fields_without_empty_overwrites(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="v2_update_owner")
    admin = _create_user(test_db_session, username="v2_update_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    account = account_resp.json()
    account_id = account["id"]

    key_resp = test_client.post(
        "/api/admin/article-distribution/api-keys",
        headers=_headers(admin),
        json={"name": "updater"},
    )
    raw_key = key_resp.json()["api_key"]

    upload_resp = test_client.post(
        "/api/v2/article-distribution/articles",
        headers={"X-API-Key": raw_key},
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "Original",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-30",
                    "theme_id": account["theme_id"],
                    "metadata": {"output_id": "260530_1"},
                }
            ],
        },
    )
    assert upload_resp.status_code == 201, upload_resp.text
    article_id = upload_resp.json()[0]["id"]
    assert upload_resp.json()[0]["project_id"] == account["project_ids"][0]
    assert upload_resp.json()[0]["metadata"] == {"output_id": "260530_1"}

    update_resp = test_client.patch(
        f"/api/v2/article-distribution/articles/{article_id}",
        headers={"X-API-Key": raw_key},
        json={
            "title": " ",
            "markdown_content": None,
            "scheduled_date": "2026-05-31",
            "project_id": 1,
            "publish_status": "published",
            "published_url": "https://example.com/articles/v2",
            "metadata": {
                "output_id": "260530_1",
                "topic": "测试选题",
                "article": {"summary": "问题和解决方案"},
            },
        },
    )
    assert update_resp.status_code == 200, update_resp.text
    updated = update_resp.json()
    assert updated["title"] == "Original"
    assert updated["markdown_content"] == "body"
    assert updated["scheduled_date"] == "2026-05-31"
    assert updated["publish_status"] == "published"
    assert updated["published_url"] == "https://example.com/articles/v2"
    assert updated["metadata"]["topic"] == "测试选题"

    empty_update_resp = test_client.patch(
        f"/api/v2/article-distribution/articles/{article_id}",
        headers={"X-API-Key": raw_key},
        json={"title": "", "metadata": {}, "published_url": ""},
    )
    assert empty_update_resp.status_code == 200
    unchanged = empty_update_resp.json()
    assert unchanged["title"] == "Original"
    assert unchanged["published_url"] == "https://example.com/articles/v2"
    assert unchanged["metadata"]["topic"] == "测试选题"

    missing_key_resp = test_client.patch(
        f"/api/v2/article-distribution/articles/{article_id}",
        json={"title": "Denied"},
    )
    assert missing_key_resp.status_code == 401


def test_inactive_accounts_are_hidden_from_directory_and_reject_uploads(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="inactive_owner")
    admin = _create_user(test_db_session, username="inactive_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "封禁号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert account_resp.status_code == 201
    account = account_resp.json()
    assert account["is_active"] is True

    inactive_resp = test_client.patch(
        f"/api/article-distribution/accounts/{account['id']}",
        headers=_headers(owner),
        json={"is_active": False},
    )
    assert inactive_resp.status_code == 200
    assert inactive_resp.json()["is_active"] is False

    key_resp = test_client.post(
        "/api/admin/article-distribution/api-keys",
        headers=_headers(admin),
        json={"name": "inactive-directory"},
    )
    assert key_resp.status_code == 201
    raw_key = key_resp.json()["api_key"]

    directory_resp = test_client.get(
        "/api/v2/article-distribution/accounts",
        headers={"X-API-Key": raw_key},
    )
    assert directory_resp.status_code == 200
    assert directory_resp.json() == []

    admin_upload_payload = {
        "account_id": account["id"],
        "articles": [
            {
                "title": "Blocked",
                "markdown_content": "body",
                "scheduled_date": "2026-05-25",
                "project_id": 1,
            }
        ],
    }
    api_upload_payload = {
        "account_id": account["id"],
        "articles": [
            {
                "title": "Blocked",
                "markdown_content": "body",
                "scheduled_date": "2026-05-25",
                "theme_id": account["theme_id"],
            }
        ],
    }
    admin_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json=admin_upload_payload,
    )
    assert admin_upload_resp.status_code == 400
    assert admin_upload_resp.json()["detail"] == "账号已停用，不能新增文章"

    api_upload_resp = test_client.post(
        "/api/v2/article-distribution/articles",
        headers={"X-API-Key": raw_key},
        json=api_upload_payload,
    )
    assert api_upload_resp.status_code == 400
    assert api_upload_resp.json()["detail"] == "账号已停用，不能新增文章"


def test_publish_status_and_filters(test_client, test_db_session: Session):
    owner = _create_user(test_db_session, username="filter_owner")
    admin = _create_user(test_db_session, username="filter_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        json={
            "account_name": "视频号",
            "platform": "shipinhao",
            "publication_type": "video",
        },
    )
    account_id = account_resp.json()["id"]
    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "Video",
                    "markdown_content": "content",
                    "scheduled_date": "2026-05-22",
                    "project_id": 1,
                }
            ],
        },
    )
    article_id = upload_resp.json()[0]["id"]

    missing_url_resp = test_client.patch(
        f"/api/article-distribution/articles/{article_id}/status",
        headers=_headers(owner),
        json={"publish_status": "published"},
    )
    assert missing_url_resp.status_code == 400

    status_resp = test_client.patch(
        f"/api/article-distribution/articles/{article_id}/status",
        headers=_headers(owner),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/video",
        },
    )
    assert status_resp.status_code == 200
    assert status_resp.json()["publish_status"] == "published"
    assert status_resp.json()["published_url"] == "https://example.com/video"

    filtered_resp = test_client.get(
        "/api/article-distribution/articles",
        headers=_headers(owner),
        params={
            "publish_status": "published",
            "publication_type": "video",
            "scheduled_from": "2026-05-22",
            "scheduled_to": "2026-05-22",
        },
    )
    assert filtered_resp.status_code == 200
    assert [item["id"] for item in filtered_resp.json()] == [article_id]

    invalid_resp = test_client.patch(
        f"/api/article-distribution/articles/{article_id}/status",
        headers=_headers(owner),
        json={"publish_status": "invalid"},
    )
    assert invalid_resp.status_code == 200
    assert invalid_resp.json()["publish_status"] == "invalid"
    assert invalid_resp.json()["published_url"] is None


def test_owner_can_add_multiple_traffic_stats_for_article(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="traffic_owner")
    admin = _create_user(test_db_session, username="traffic_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        json={
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert account_resp.status_code == 201
    account_id = account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "Traffic Article",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-24",
                "project_id": 1,
                }
            ],
        },
    )
    assert upload_resp.status_code == 201
    article_id = upload_resp.json()[0]["id"]

    first_resp = test_client.post(
        f"/api/article-distribution/articles/{article_id}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 50,
            "like_count": 5,
            "favorite_count": 2,
            "share_count": 1,
        },
    )
    assert first_resp.status_code == 201, first_resp.text
    assert first_resp.json()["recorded_at"] is not None

    second_resp = test_client.post(
        f"/api/article-distribution/articles/{article_id}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 180,
            "like_count": 18,
            "favorite_count": 9,
            "share_count": 4,
            "recorded_at": "2099-05-24T10:00:00+00:00",
        },
    )
    assert second_resp.status_code == 201, second_resp.text
    assert second_resp.json()["read_count"] == 180
    assert second_resp.json()["recorded_at"].startswith("2099-05-24T10:00:00")

    list_resp = test_client.get(
        f"/api/article-distribution/articles/{article_id}/traffic-stats",
        headers=_headers(owner),
    )
    assert list_resp.status_code == 200
    assert [item["read_count"] for item in list_resp.json()] == [180, 50]

    summary_resp = test_client.get(
        "/api/article-distribution/traffic-stats/articles/page",
        headers=_headers(owner),
    )
    assert summary_resp.status_code == 200
    summary = summary_resp.json()
    assert summary["total"] == 1
    assert summary["items"][0]["article"]["id"] == article_id
    assert summary["items"][0]["latest_stat"]["read_count"] == 180
    assert summary["items"][0]["record_count"] == 2

    report_resp = test_client.get(
        "/api/article-distribution/reports/unpublished",
        headers=_headers(admin),
    )
    assert report_resp.status_code == 200
    assert report_resp.json()["summary"]["read_count"] == 180
    assert report_resp.json()["summary"]["like_count"] == 18
    assert report_resp.json()["users"][0]["read_count"] == 180
    assert report_resp.json()["users"][0]["articles"] == []
    detail_resp = test_client.get(
        f"/api/article-distribution/reports/unpublished/users/{owner.id}",
        headers=_headers(admin),
    )
    assert detail_resp.status_code == 200
    report_article = detail_resp.json()["articles"][0]
    assert report_article["id"] == article_id
    assert report_article["latest_traffic_stat"]["read_count"] == 180
    assert report_article["latest_traffic_stat"]["like_count"] == 18


def test_metadata_dashboard_groups_articles_by_output_id_and_requires_scope(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="metadata_owner")
    viewer = _create_user(test_db_session, username="metadata_viewer")
    admin = _create_user(test_db_session, username="metadata_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    account_id = account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "Main",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-30",
                    "project_id": 1,
                    "metadata": {
                        "output_id": "260530_5",
                        "topic": "AIFC Douju算力预算控制",
                        "article": {
                            "role": "main",
                            "summary": "拆解算力预算痛点并给出积分控制方案",
                            "materials_used": [
                                {"title": "首届 Douju 全球 AI 互动影视创作大赛细则"},
                                {"title": "AI短剧风口下：凌晨五点，抢算力的人"},
                            ],
                        },
                    },
                },
                {
                    "title": "Variant",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-31",
                    "project_id": 1,
                    "metadata": {
                        "output_id": "260530_5",
                        "angle_label": "案例复盘型",
                        "audience_label": "案例复盘型",
                        "article": {
                            "role": "variant",
                            "summary": "用案例复盘承接预算方案",
                            "based_on_output_id": "260530_5",
                        },
                    },
                },
                {
                    "title": "No metadata",
                    "markdown_content": "body",
                    "scheduled_date": "2026-06-01",
                "project_id": 1,
                },
            ],
        },
    )
    assert upload_resp.status_code == 201, upload_resp.text
    main_id = upload_resp.json()[0]["id"]
    variant_id = upload_resp.json()[1]["id"]

    traffic_resp = test_client.post(
        f"/api/article-distribution/articles/{main_id}/traffic-stats",
        headers=_headers(owner),
        json={"read_count": 300, "like_count": 30, "favorite_count": 12},
    )
    assert traffic_resp.status_code == 201

    denied_resp = test_client.get(
        "/api/article-distribution/reports/metadata-dashboard",
        headers=_headers(viewer),
    )
    assert denied_resp.status_code == 403

    dashboard_resp = test_client.get(
        "/api/article-distribution/reports/metadata-dashboard",
        headers=_headers(admin),
        params={
            "scheduled_from": "2026-05-30",
            "scheduled_to": "2026-05-31",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert dashboard_resp.status_code == 200, dashboard_resp.text
    dashboard = dashboard_resp.json()
    assert dashboard["summary"]["topic_count"] == 1
    assert dashboard["summary"]["article_count"] == 2
    assert dashboard["summary"]["material_count"] == 2
    assert dashboard["summary"]["read_count"] == 300
    assert dashboard["total"] == 2

    topic = dashboard["topics"][0]
    assert topic["output_id"] == "260530_5"
    assert topic["topic"] == "AIFC Douju算力预算控制"
    assert topic["materials"] == [
        "首届 Douju 全球 AI 互动影视创作大赛细则",
        "AI短剧风口下：凌晨五点，抢算力的人",
    ]
    assert [article["id"] for article in topic["articles"]] == [variant_id, main_id]
    assert topic["articles"][0]["article_role"] == "variant"
    assert topic["articles"][0]["angle_label"] == "案例复盘型"
    assert topic["articles"][1]["summary"] == "拆解算力预算痛点并给出积分控制方案"
    assert topic["articles"][1]["latest_traffic_stat"]["read_count"] == 300

    fallback_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "备用号",
            "platform": "fallback",
            "publication_type": "article",
        },
    )
    fallback_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": fallback_account_resp.json()["id"],
            "articles": [
                {
                    "title": "Fallback only",
                    "markdown_content": "body",
                    "scheduled_date": "2026-06-02",
                "project_id": 1,
                }
            ],
        },
    )
    assert fallback_upload_resp.status_code == 201
    fallback_resp = test_client.get(
        "/api/article-distribution/reports/metadata-dashboard",
        headers=_headers(admin),
        params={"platform": "fallback"},
    )
    assert fallback_resp.status_code == 200
    fallback_dashboard = fallback_resp.json()
    assert fallback_dashboard["total"] == 1
    assert fallback_dashboard["topics"][0]["topic"] == "未设置选题"
    assert fallback_dashboard["topics"][0]["articles"][0]["title"] == "Fallback only"


def test_user_cannot_manage_other_users_traffic_stats(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="traffic_private_owner")
    other = _create_user(test_db_session, username="traffic_private_other")
    admin = _create_user(
        test_db_session, username="traffic_private_admin", role=UserRole.ADMIN
    )

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(owner),
        json={
            "account_name": "主号",
            "platform": "zhihu",
            "publication_type": "article",
        },
    )
    account_id = account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "Private Traffic",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-24",
                "project_id": 1,
                }
            ],
        },
    )
    article_id = upload_resp.json()[0]["id"]

    denied_create_resp = test_client.post(
        f"/api/article-distribution/articles/{article_id}/traffic-stats",
        headers=_headers(other),
        json={"read_count": 1},
    )
    assert denied_create_resp.status_code == 403

    owner_create_resp = test_client.post(
        f"/api/article-distribution/articles/{article_id}/traffic-stats",
        headers=_headers(owner),
        json={"read_count": 1},
    )
    stat_id = owner_create_resp.json()["id"]

    denied_list_resp = test_client.get(
        f"/api/article-distribution/articles/{article_id}/traffic-stats",
        headers=_headers(other),
    )
    assert denied_list_resp.status_code == 403

    denied_delete_resp = test_client.delete(
        f"/api/article-distribution/traffic-stats/{stat_id}",
        headers=_headers(other),
    )
    assert denied_delete_resp.status_code == 403


def test_missing_traffic_report_lists_published_articles_without_today_stats(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="missing_traffic_owner")
    viewer = _create_user(test_db_session, username="missing_traffic_viewer")
    admin = _create_user(
        test_db_session, username="missing_traffic_admin", role=UserRole.ADMIN
    )

    active_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    inactive_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "停用号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert active_account_resp.status_code == 201
    assert inactive_account_resp.status_code == 201
    active_account_id = active_account_resp.json()["id"]
    inactive_account_id = inactive_account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": active_account_id,
            "articles": [
                {
                    "title": "No stat today",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-20",
                "project_id": 1,
                },
                {
                    "title": "Has stat today",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-21",
                "project_id": 1,
                },
                {
                    "title": "Only yesterday stat",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-22",
                "project_id": 1,
                },
                {
                    "title": "Unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-23",
                "project_id": 1,
                },
                {
                    "title": "Invalid",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-24",
                "project_id": 1,
                },
                {
                    "title": "Published without url",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-25",
                "project_id": 1,
                },
            ],
        },
    )
    inactive_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": inactive_account_id,
            "articles": [
                {
                    "title": "Inactive no stat today",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-26",
                "project_id": 1,
                }
            ],
        },
    )
    assert upload_resp.status_code == 201
    assert inactive_upload_resp.status_code == 201
    articles = upload_resp.json()
    inactive_article_id = inactive_upload_resp.json()[0]["id"]

    published_ids = [item["id"] for item in articles[:3]]
    invalid_id = articles[4]["id"]
    no_url_id = articles[5]["id"]
    for article_id in [*published_ids, inactive_article_id]:
        publish_resp = test_client.patch(
            f"/api/article-distribution/articles/{article_id}/status",
            headers=_headers(owner),
            json={
                "publish_status": "published",
                "published_url": f"https://example.com/articles/{article_id}",
            },
        )
        assert publish_resp.status_code == 200

    invalid_resp = test_client.patch(
        f"/api/article-distribution/articles/{invalid_id}/status",
        headers=_headers(owner),
        json={"publish_status": "invalid"},
    )
    assert invalid_resp.status_code == 200

    no_url_article = (
        test_db_session.query(ArticleDistributionArticle)
        .filter(ArticleDistributionArticle.id == no_url_id)
        .one()
    )
    no_url_article.publish_status = "published"
    no_url_article.published_url = None
    test_db_session.commit()

    today_stat_resp = test_client.post(
        f"/api/article-distribution/articles/{published_ids[1]}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 100,
            "like_count": 10,
            "recorded_at": "2026-05-28T10:00:00+00:00",
        },
    )
    yesterday_stat_resp = test_client.post(
        f"/api/article-distribution/articles/{published_ids[2]}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 50,
            "like_count": 5,
            "recorded_at": "2026-05-27T23:00:00+00:00",
        },
    )
    assert today_stat_resp.status_code == 201
    assert yesterday_stat_resp.status_code == 201

    deactivate_resp = test_client.patch(
        f"/api/article-distribution/accounts/{inactive_account_id}",
        headers=_headers(owner),
        json={"is_active": False},
    )
    assert deactivate_resp.status_code == 200

    params = {
        "recorded_from": "2026-05-28T00:00:00+00:00",
        "recorded_to": "2026-05-29T00:00:00+00:00",
        "page": 1,
        "page_size": 10,
    }
    denied_resp = test_client.get(
        "/api/article-distribution/reports/missing-traffic",
        headers=_headers(viewer),
        params=params,
    )
    assert denied_resp.status_code == 403

    viewer.scope_overrides = auth_service.serialize_scopes(
        [
            *auth_service.get_role_scopes(UserRole.USER),
            auth_service.SCOPE_ARTICLE_DISTRIBUTION_REPORT_READ,
        ]
    )
    test_db_session.commit()
    test_db_session.refresh(viewer)

    report_resp = test_client.get(
        "/api/article-distribution/reports/missing-traffic",
        headers=_headers(viewer),
        params=params,
    )
    assert report_resp.status_code == 200, report_resp.text
    report = report_resp.json()
    assert report["total"] == 1
    assert [item["title"] for item in report["items"]] == ["Only yesterday stat"]
    assert report["items"][0]["latest_traffic_stat"]["read_count"] == 50
    assert all(item["published_url"] for item in report["items"])

    user_report_resp = test_client.get(
        "/api/article-distribution/reports/missing-traffic/users",
        headers=_headers(viewer),
        params=params,
    )
    assert user_report_resp.status_code == 200, user_report_resp.text
    user_report = user_report_resp.json()
    assert user_report["summary"] == {
        "total_users": 1,
        "missing_articles": 1,
        "read_count": 50,
        "like_count": 5,
        "favorite_count": 0,
        "share_count": 0,
    }
    assert user_report["users"][0]["user_id"] == owner.id
    assert user_report["users"][0]["missing_count"] == 1
    assert user_report["users"][0]["articles"] == []

    detail_resp = test_client.get(
        f"/api/article-distribution/reports/missing-traffic/users/{owner.id}",
        headers=_headers(viewer),
        params=params,
    )
    assert detail_resp.status_code == 200, detail_resp.text
    detail = detail_resp.json()
    assert detail["missing_count"] == 1
    assert [item["title"] for item in detail["articles"]] == ["Only yesterday stat"]

    all_accounts_resp = test_client.get(
        "/api/article-distribution/reports/missing-traffic",
        headers=_headers(admin),
        params={**params, "account_status": "all"},
    )
    assert all_accounts_resp.status_code == 200
    assert all_accounts_resp.json()["total"] == 2
    assert all_accounts_resp.json()["items"][0]["title"] == "Inactive no stat today"

    filtered_resp = test_client.get(
        "/api/article-distribution/reports/missing-traffic",
        headers=_headers(admin),
        params={
            **params,
            "scheduled_from": "2026-05-22",
            "scheduled_to": "2026-05-22",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert filtered_resp.status_code == 200
    assert filtered_resp.json()["total"] == 1
    assert filtered_resp.json()["items"][0]["title"] == "Only yesterday stat"

    invalid_range_resp = test_client.get(
        "/api/article-distribution/reports/missing-traffic",
        headers=_headers(admin),
        params={
            "recorded_from": "2026-05-29T00:00:00+00:00",
            "recorded_to": "2026-05-28T00:00:00+00:00",
        },
    )
    assert invalid_range_resp.status_code == 400


def test_report_overview_supports_views_filters_and_topic_permission(
    test_client, test_db_session: Session
):
    owner = _create_user(
        test_db_session,
        username="overview_owner",
        name="Overview",
        wechat_nickname="概览微信",
        wechat_id="overview_wx",
    )
    viewer = _create_user(test_db_session, username="overview_viewer")
    admin = _create_user(test_db_session, username="overview_admin", role=UserRole.ADMIN)
    default_project = bootstrap_default_project_theme(test_db_session)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert account_resp.status_code == 201
    account_theme_id = account_resp.json()["theme_id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_resp.json()["id"],
            "articles": [
                {
                    "title": "Overview missing stat",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-20",
                    "project_id": default_project.id,
                    "metadata": {"output_id": "overview_topic", "topic": "统一报表"},
                },
                {
                    "title": "Overview unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-21",
                    "project_id": default_project.id,
                },
            ],
        },
    )
    assert upload_resp.status_code == 201
    published_id = upload_resp.json()[0]["id"]

    publish_resp = test_client.patch(
        f"/api/article-distribution/articles/{published_id}/status",
        headers=_headers(owner),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/overview",
        },
    )
    assert publish_resp.status_code == 200

    stat_resp = test_client.post(
        f"/api/article-distribution/articles/{published_id}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 50,
            "like_count": 5,
            "recorded_at": "2026-05-27T23:00:00+00:00",
        },
    )
    assert stat_resp.status_code == 201

    denied_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
    )
    assert denied_resp.status_code == 403

    viewer.scope_overrides = auth_service.serialize_scopes(
        [
            *auth_service.get_role_scopes(UserRole.USER),
            auth_service.SCOPE_ARTICLE_DISTRIBUTION_REPORT_READ,
        ]
    )
    test_db_session.commit()
    test_db_session.refresh(viewer)

    users_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={"view": "users", "keyword": "Overview"},
    )
    assert users_resp.status_code == 200, users_resp.text
    users_report = users_resp.json()
    assert users_report["summary"]["total_articles"] == 2
    assert users_report["summary"]["published_articles"] == 1
    assert users_report["summary"]["unpublished_articles"] == 1
    assert users_report["items"][0]["item_type"] == "user"
    assert users_report["items"][0]["wechat_nickname"] == "概览微信"
    assert users_report["items"][0]["wechat_id"] == "overview_wx"
    assert users_report["items"][0]["published_count"] == 1
    assert users_report["items"][0]["remaining_count"] == 1
    assert users_report["items"][0]["articles"] == []

    project_filtered_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={"view": "users", "project_id": default_project.id},
    )
    assert project_filtered_resp.status_code == 200
    assert project_filtered_resp.json()["summary"]["total_articles"] == 2

    theme_filtered_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={"view": "users", "theme_id": account_theme_id},
    )
    assert theme_filtered_resp.status_code == 200
    assert theme_filtered_resp.json()["summary"]["total_articles"] == 2

    empty_filter_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={"view": "users", "project_id": default_project.id + 10000},
    )
    assert empty_filter_resp.status_code == 200
    assert empty_filter_resp.json()["summary"]["total_articles"] == 0

    user_articles_resp = test_client.get(
        "/api/article-distribution/reports/overview/articles",
        headers=_headers(viewer),
        params={
            "user_id": owner.id,
            "keyword": "Overview",
            "project_id": default_project.id,
            "theme_id": account_theme_id,
            "page": 1,
            "page_size": 1,
        },
    )
    assert user_articles_resp.status_code == 200, user_articles_resp.text
    user_articles = user_articles_resp.json()
    assert user_articles["total"] == 2
    assert user_articles["page_size"] == 1
    assert len(user_articles["items"]) == 1
    assert user_articles["items"][0]["item_type"] == "article"
    assert user_articles["items"][0]["wechat_nickname"] == "概览微信"
    assert user_articles["items"][0]["wechat_id"] == "overview_wx"
    assert "markdown_content" not in user_articles["items"][0]
    assert "metadata" not in user_articles["items"][0]
    assert "summary" not in user_articles["items"][0]

    detail_resp = test_client.get(
        f"/api/article-distribution/reports/overview/articles/{published_id}",
        headers=_headers(viewer),
        params={
            "recorded_from": "2026-05-28T00:00:00+00:00",
            "recorded_to": "2026-05-29T00:00:00+00:00",
        },
    )
    assert detail_resp.status_code == 200, detail_resp.text
    detail = detail_resp.json()
    assert detail["markdown_content"] == "body"
    assert detail["metadata"] == {"output_id": "overview_topic", "topic": "统一报表"}
    assert detail["wechat_nickname"] == "概览微信"
    assert detail["wechat_id"] == "overview_wx"
    assert detail["missing_traffic"] is True

    missing_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={
            "view": "articles",
            "missing_traffic_only": True,
            "recorded_from": "2026-05-28T00:00:00+00:00",
            "recorded_to": "2026-05-29T00:00:00+00:00",
        },
    )
    assert missing_resp.status_code == 200, missing_resp.text
    missing_report = missing_resp.json()
    assert missing_report["total"] == 1
    assert missing_report["summary"]["missing_articles"] == 1
    assert missing_report["items"][0]["item_type"] == "article"
    assert missing_report["items"][0]["missing_traffic"] is True
    assert missing_report["items"][0]["latest_traffic_stat"]["read_count"] == 50

    missing_articles_resp = test_client.get(
        "/api/article-distribution/reports/overview/articles",
        headers=_headers(viewer),
        params={
            "missing_traffic_only": True,
            "recorded_from": "2026-05-28T00:00:00+00:00",
            "recorded_to": "2026-05-29T00:00:00+00:00",
        },
    )
    assert missing_articles_resp.status_code == 200, missing_articles_resp.text
    assert missing_articles_resp.json()["total"] == 1

    invalid_missing_articles_resp = test_client.get(
        "/api/article-distribution/reports/overview/articles",
        headers=_headers(viewer),
        params={"missing_traffic_only": True},
    )
    assert invalid_missing_articles_resp.status_code == 400

    topic_denied_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={"view": "topics"},
    )
    assert topic_denied_resp.status_code == 403

    topic_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(admin),
        params={"view": "topics"},
    )
    assert topic_resp.status_code == 200, topic_resp.text
    topic_report = topic_resp.json()
    assert topic_report["items"][0]["item_type"] == "topic"
    topic_item = next(item for item in topic_report["items"] if item["topic"] == "统一报表")
    assert topic_item["articles"] == []

    topic_articles_resp = test_client.get(
        "/api/article-distribution/reports/overview/articles",
        headers=_headers(admin),
        params={"topic_key": topic_item["key"]},
    )
    assert topic_articles_resp.status_code == 200, topic_articles_resp.text
    topic_articles = topic_articles_resp.json()
    assert topic_articles["total"] == 1
    assert topic_articles["items"][0]["title"] == "Overview missing stat"


def test_report_overview_sorts_metrics_before_pagination(
    test_client, test_db_session: Session
):
    owner_a = _create_user(test_db_session, username="overview_sort_a", name="Sort A")
    owner_b = _create_user(test_db_session, username="overview_sort_b", name="Sort B")
    viewer = _create_user(test_db_session, username="overview_sort_viewer")
    admin = _create_user(
        test_db_session, username="overview_sort_admin", role=UserRole.ADMIN
    )
    viewer.scope_overrides = auth_service.serialize_scopes(
        [
            *auth_service.get_role_scopes(UserRole.USER),
            auth_service.SCOPE_ARTICLE_DISTRIBUTION_REPORT_READ,
        ]
    )
    test_db_session.commit()
    test_db_session.refresh(viewer)

    article_ids: list[int] = []
    for owner, title, read_count, like_count in [
        (owner_a, "Metric Low", 20, 30),
        (owner_b, "Metric High", 200, 3),
    ]:
        account_resp = test_client.post(
            "/api/article-distribution/accounts",
            headers=_headers(admin),
            json={
                "user_id": owner.id,
                "account_name": f"{title} Account",
                "platform": "wechat",
                "publication_type": "article",
            },
        )
        assert account_resp.status_code == 201
        upload_resp = test_client.post(
            "/api/admin/article-distribution/articles",
            headers=_headers(admin),
            json={
                "account_id": account_resp.json()["id"],
                "articles": [
                    {
                        "title": title,
                        "markdown_content": "body",
                        "scheduled_date": "2026-05-20",
                    "project_id": 1,
                    },
                ],
            },
        )
        assert upload_resp.status_code == 201
        article_id = upload_resp.json()[0]["id"]
        article_ids.append(article_id)
        stat_resp = test_client.post(
            f"/api/article-distribution/articles/{article_id}/traffic-stats",
            headers=_headers(owner),
            json={
                "read_count": read_count,
                "like_count": like_count,
                "recorded_at": "2026-05-28T10:00:00+00:00",
            },
        )
        assert stat_resp.status_code == 201, stat_resp.text

    users_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={
            "view": "users",
            "sort_by": "read_count",
            "sort_order": "desc",
            "page": 2,
            "page_size": 1,
        },
    )
    assert users_resp.status_code == 200, users_resp.text
    assert users_resp.json()["total"] == 2
    assert users_resp.json()["items"][0]["username"] == owner_a.username

    articles_resp = test_client.get(
        "/api/article-distribution/reports/overview",
        headers=_headers(viewer),
        params={
            "view": "articles",
            "sort_by": "like_count",
            "sort_order": "asc",
            "page": 1,
            "page_size": 1,
        },
    )
    assert articles_resp.status_code == 200, articles_resp.text
    assert articles_resp.json()["items"][0]["title"] == "Metric High"

    detail_resp = test_client.get(
        "/api/article-distribution/reports/overview/articles",
        headers=_headers(viewer),
        params={
            "sort_by": "read_count",
            "sort_order": "desc",
            "page": 1,
            "page_size": 1,
        },
    )
    assert detail_resp.status_code == 200, detail_resp.text
    assert detail_resp.json()["items"][0]["id"] == article_ids[1]


def test_report_overview_export_supports_csv_xlsx_and_permissions(
    test_client, test_db_session: Session
):
    owner = _create_user(
        test_db_session,
        username="overview_export_owner",
        wechat_nickname="导出微信",
        wechat_id="export_wx",
    )
    viewer = _create_user(test_db_session, username="overview_export_viewer")
    admin = _create_user(
        test_db_session, username="overview_export_admin", role=UserRole.ADMIN
    )

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "导出公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert account_resp.status_code == 201

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_resp.json()["id"],
            "articles": [
                {
                    "title": "Export published",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-20",
                    "project_id": 1,
                    "metadata": {
                        "output_id": "export_topic",
                        "topic": "导出选题",
                        "article": {"role": "main"},
                    },
                },
                {
                    "title": "Export unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-21",
                "project_id": 1,
                },
            ],
        },
    )
    assert upload_resp.status_code == 201
    published_id = upload_resp.json()[0]["id"]

    publish_resp = test_client.patch(
        f"/api/article-distribution/articles/{published_id}/status",
        headers=_headers(owner),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/export",
        },
    )
    assert publish_resp.status_code == 200

    stat_resp = test_client.post(
        f"/api/article-distribution/articles/{published_id}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 88,
            "like_count": 8,
            "favorite_count": 6,
            "share_count": 4,
            "recorded_at": "2026-05-27T10:00:00+00:00",
        },
    )
    assert stat_resp.status_code == 201

    denied_resp = test_client.get(
        "/api/article-distribution/reports/overview/export",
        headers=_headers(viewer),
    )
    assert denied_resp.status_code == 403

    viewer.scope_overrides = auth_service.serialize_scopes(
        [
            *auth_service.get_role_scopes(UserRole.USER),
            auth_service.SCOPE_ARTICLE_DISTRIBUTION_REPORT_READ,
        ]
    )
    test_db_session.commit()
    test_db_session.refresh(viewer)

    csv_resp = test_client.get(
        "/api/article-distribution/reports/overview/export",
        headers=_headers(viewer),
        params={
            "view": "articles",
            "format": "csv",
            "keyword": "Export",
            "scheduled_from": "2026-05-20",
            "scheduled_to": "2026-05-21",
        },
    )
    assert csv_resp.status_code == 200, csv_resp.text
    assert csv_resp.headers["content-type"].startswith("text/csv")
    assert "overview-articles-2026-05-21.csv" in csv_resp.headers[
        "content-disposition"
    ]
    csv_rows = list(csv.DictReader(StringIO(csv_resp.content.decode("utf-8-sig"))))
    assert [row["标题"] for row in csv_rows] == [
        "Export unpublished",
        "Export published",
    ]
    published_row = csv_rows[1]
    assert published_row["微信昵称"] == "导出微信"
    assert published_row["微信号"] == "export_wx"
    assert published_row["发布账号"] == "导出公众号"
    assert published_row["发布状态"] == "已发布"
    assert published_row["发布链接"] == "https://example.com/export"
    assert published_row["阅读量"] == "88"
    assert published_row["点赞量"] == "8"
    assert published_row["选题"] == "导出选题"

    xlsx_resp = test_client.get(
        "/api/article-distribution/reports/overview/export",
        headers=_headers(viewer),
        params={"view": "users", "format": "xlsx", "keyword": "Export"},
    )
    assert xlsx_resp.status_code == 200, xlsx_resp.text
    assert xlsx_resp.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(xlsx_resp.content), read_only=True)
    worksheet = workbook["用户汇总"]
    rows = list(worksheet.iter_rows(values_only=True))
    assert rows[0][:7] == (
        "用户ID",
        "负责人",
        "用户名",
        "微信昵称",
        "微信号",
        "邮箱",
        "剩余未发布",
    )
    assert rows[1][0] == owner.id
    assert rows[1][3] == "导出微信"
    assert rows[1][4] == "export_wx"
    assert rows[1][7] == 1
    assert rows[1][11] == 88
    workbook.close()

    topic_denied_resp = test_client.get(
        "/api/article-distribution/reports/overview/export",
        headers=_headers(viewer),
        params={"view": "topics"},
    )
    assert topic_denied_resp.status_code == 403

    topic_resp = test_client.get(
        "/api/article-distribution/reports/overview/export",
        headers=_headers(admin),
        params={"view": "topics", "format": "csv"},
    )
    assert topic_resp.status_code == 200, topic_resp.text
    topic_rows = list(csv.DictReader(StringIO(topic_resp.content.decode("utf-8-sig"))))
    topic_row = next(row for row in topic_rows if row["Output ID"] == "export_topic")
    assert topic_row["选题"] == "导出选题"

    invalid_missing_resp = test_client.get(
        "/api/article-distribution/reports/overview/export",
        headers=_headers(viewer),
        params={"view": "articles", "missing_traffic_only": True},
    )
    assert invalid_missing_resp.status_code == 400


def test_admin_can_export_publicity_records_csv(test_client, test_db_session: Session):
    owner = _create_user(test_db_session, username="publicity_owner", name="Owner Name")
    viewer = _create_user(test_db_session, username="publicity_viewer")
    admin = _create_user(
        test_db_session, username="publicity_admin", role=UserRole.ADMIN
    )
    project_dao = ProjectManagementDAO(test_db_session)
    default_project = bootstrap_default_project_theme(test_db_session)
    default_theme = project_dao.get_theme_by_name("AI")
    assert default_theme is not None
    other_theme = project_dao.create_theme(Theme(name="Healthcare", is_active=True))
    other_project = project_dao.create_project(
        Project(name="MED", code="MEDPROJA", is_active=True)
    )
    project_dao.replace_project_themes(other_project.id, [other_theme.id])
    project_dao.add_user_project(owner.id, other_project.id)

    active_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
            "theme_id": default_theme.id,
        },
    )
    inactive_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "旧号",
            "platform": "wechat",
            "publication_type": "article",
            "theme_id": default_theme.id,
        },
    )
    other_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "知乎号",
            "platform": "zhihu",
            "publication_type": "article",
            "theme_id": other_theme.id,
        },
    )
    assert active_account_resp.status_code == 201
    assert inactive_account_resp.status_code == 201
    assert other_account_resp.status_code == 201
    active_account_id = active_account_resp.json()["id"]
    inactive_account_id = inactive_account_resp.json()["id"]
    other_account_id = other_account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": active_account_id,
            "articles": [
                {
                    "title": "Published A",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-27",
                    "project_id": default_project.id,
                },
                {
                    "title": "Unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-27",
                    "project_id": default_project.id,
                },
                {
                    "title": "Invalid",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-27",
                    "project_id": default_project.id,
                },
                {
                    "title": "No URL",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-27",
                    "project_id": default_project.id,
                },
                {
                    "title": "Future",
                    "markdown_content": "body",
                    "scheduled_date": "2099-01-01",
                    "project_id": default_project.id,
                },
            ],
        },
    )
    inactive_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": inactive_account_id,
            "articles": [
                {
                    "title": "Inactive published",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-26",
                    "project_id": default_project.id,
                }
            ],
        },
    )
    other_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": other_account_id,
            "articles": [
                {
                    "title": "Zhihu published",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-27",
                    "project_id": other_project.id,
                }
            ],
        },
    )
    assert upload_resp.status_code == 201
    assert inactive_upload_resp.status_code == 201
    assert other_upload_resp.status_code == 201

    articles = upload_resp.json()
    published_id = articles[0]["id"]
    invalid_id = articles[2]["id"]
    no_url_id = articles[3]["id"]
    future_id = articles[4]["id"]
    inactive_id = inactive_upload_resp.json()[0]["id"]
    other_id = other_upload_resp.json()[0]["id"]
    for article_id in [published_id, future_id, inactive_id, other_id]:
        publish_resp = test_client.patch(
            f"/api/article-distribution/articles/{article_id}/status",
            headers=_headers(owner),
            json={
                "publish_status": "published",
                "published_url": f"https://example.com/articles/{article_id}",
            },
        )
        assert publish_resp.status_code == 200

    invalid_resp = test_client.patch(
        f"/api/article-distribution/articles/{invalid_id}/status",
        headers=_headers(owner),
        json={"publish_status": "invalid"},
    )
    assert invalid_resp.status_code == 200

    no_url_article = (
        test_db_session.query(ArticleDistributionArticle)
        .filter(ArticleDistributionArticle.id == no_url_id)
        .one()
    )
    no_url_article.publish_status = "published"
    no_url_article.published_url = None
    test_db_session.commit()

    first_stat_resp = test_client.post(
        f"/api/article-distribution/articles/{published_id}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 10,
            "like_count": 1,
            "recorded_at": "2026-05-27T10:00:00+00:00",
        },
    )
    latest_stat_resp = test_client.post(
        f"/api/article-distribution/articles/{published_id}/traffic-stats",
        headers=_headers(owner),
        json={
            "read_count": 20,
            "like_count": 2,
            "favorite_count": 3,
            "share_count": 4,
            "recorded_at": "2026-05-28T10:00:00+00:00",
        },
    )
    assert first_stat_resp.status_code == 201
    assert latest_stat_resp.status_code == 201

    deactivate_resp = test_client.patch(
        f"/api/article-distribution/accounts/{inactive_account_id}",
        headers=_headers(owner),
        json={"is_active": False},
    )
    assert deactivate_resp.status_code == 200

    denied_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(viewer),
    )
    assert denied_resp.status_code == 403

    export_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
    )
    assert export_resp.status_code == 200, export_resp.text
    assert export_resp.headers["content-type"].startswith("text/csv")
    assert "publicity-records-" in export_resp.headers["content-disposition"]
    rows = list(csv.DictReader(StringIO(export_resp.content.decode("utf-8-sig"))))

    assert [row["标题"] for row in rows] == [
        "Published A",
        "Zhihu published",
        "Inactive published",
    ]
    published_row = rows[0]
    assert published_row["负责人"] == "Owner Name"
    assert published_row["平台"] == "wechat"
    assert published_row["发布账号"] == "公众号"
    assert published_row["发布类型"] == "文章"
    assert published_row["账号状态"] == "启用"
    assert published_row["链接"] == f"https://example.com/articles/{published_id}"
    assert published_row["最近阅读量"] == "20"
    assert published_row["最近点赞量"] == "2"
    assert published_row["最近收藏量"] == "3"
    assert published_row["最近转发量"] == "4"
    assert published_row["最近统计时间"].startswith("2026-05-28T10:00:00")

    active_filtered_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
        params={
            "scheduled_from": "2026-05-27",
            "scheduled_to": "2026-05-27",
            "platform": "wechat",
            "publication_type": "article",
            "account_status": "active",
        },
    )
    assert active_filtered_resp.status_code == 200
    filtered_rows = list(
        csv.DictReader(StringIO(active_filtered_resp.content.decode("utf-8-sig")))
    )
    assert [row["标题"] for row in filtered_rows] == ["Published A"]

    project_filtered_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
        params={"project_id": other_project.id, "scheduled_to": "2026-05-28"},
    )
    assert project_filtered_resp.status_code == 200
    project_rows = list(
        csv.DictReader(StringIO(project_filtered_resp.content.decode("utf-8-sig")))
    )
    assert [row["标题"] for row in project_rows] == ["Zhihu published"]

    theme_filtered_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
        params={"theme_id": other_theme.id, "scheduled_to": "2026-05-28"},
    )
    assert theme_filtered_resp.status_code == 200
    theme_rows = list(
        csv.DictReader(StringIO(theme_filtered_resp.content.decode("utf-8-sig")))
    )
    assert [row["标题"] for row in theme_rows] == ["Zhihu published"]

    mismatched_filter_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
        params={
            "project_id": default_project.id,
            "theme_id": other_theme.id,
            "scheduled_to": "2026-05-28",
        },
    )
    assert mismatched_filter_resp.status_code == 200
    mismatched_rows = list(
        csv.DictReader(StringIO(mismatched_filter_resp.content.decode("utf-8-sig")))
    )
    assert mismatched_rows == []

    inactive_filtered_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
        params={"account_status": "inactive", "scheduled_to": "2026-05-28"},
    )
    assert inactive_filtered_resp.status_code == 200
    inactive_rows = list(
        csv.DictReader(StringIO(inactive_filtered_resp.content.decode("utf-8-sig")))
    )
    assert [row["标题"] for row in inactive_rows] == ["Inactive published"]

    invalid_range_resp = test_client.get(
        "/api/admin/article-distribution/publicity-records.csv",
        headers=_headers(admin),
        params={"scheduled_from": "2026-05-29", "scheduled_to": "2026-05-28"},
    )
    assert invalid_range_resp.status_code == 400


def test_admin_can_list_update_and_delete_all_articles(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="admin_crud_owner")
    admin = _create_user(test_db_session, username="admin_crud_admin", role=UserRole.ADMIN)

    account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert account_resp.status_code == 201
    account_id = account_resp.json()["id"]

    upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_id,
            "articles": [
                {
                    "title": "Original",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-23",
                "project_id": 1,
                }
            ],
        },
    )
    assert upload_resp.status_code == 201
    article_id = upload_resp.json()[0]["id"]

    admin_list_resp = test_client.get(
        "/api/article-distribution/articles", headers=_headers(admin)
    )
    assert admin_list_resp.status_code == 200
    assert [item["id"] for item in admin_list_resp.json()] == [article_id]

    update_resp = test_client.patch(
        f"/api/admin/article-distribution/articles/{article_id}",
        headers=_headers(admin),
        json={
            "title": "Updated",
            "markdown_content": "updated body",
            "scheduled_date": "2026-05-24",
        "project_id": 1,
        },
    )
    assert update_resp.status_code == 200, update_resp.text
    assert update_resp.json()["title"] == "Updated"
    assert update_resp.json()["markdown_content"] == "updated body"
    assert update_resp.json()["scheduled_date"] == "2026-05-24"

    delete_resp = test_client.delete(
        f"/api/admin/article-distribution/articles/{article_id}",
        headers=_headers(admin),
    )
    assert delete_resp.status_code == 204

    owner_list_resp = test_client.get(
        "/api/article-distribution/articles", headers=_headers(owner)
    )
    assert owner_list_resp.status_code == 200
    assert owner_list_resp.json() == []


def test_paginated_article_list_supports_filters_counts_and_access_scope(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="page_owner")
    other = _create_user(test_db_session, username="page_other")
    admin = _create_user(test_db_session, username="page_admin", role=UserRole.ADMIN)

    owner_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "Owner Account",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    other_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": other.id,
            "account_name": "Other Account",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    assert owner_account_resp.status_code == 201
    assert other_account_resp.status_code == 201
    owner_account_id = owner_account_resp.json()["id"]
    other_account_id = other_account_resp.json()["id"]

    upload_owner_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": owner_account_id,
            "articles": [
                {
                    "title": f"Owner Article {index}",
                    "markdown_content": "body",
                    "scheduled_date": f"2026-06-{index:02d}",
                    "project_id": 1,
                }
                for index in range(1, 13)
            ],
        },
    )
    upload_other_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": other_account_id,
            "articles": [
                {
                    "title": "Other Article",
                    "markdown_content": "body",
                    "scheduled_date": "2026-06-20",
                    "project_id": 1,
                }
            ],
        },
    )
    assert upload_owner_resp.status_code == 201
    assert upload_other_resp.status_code == 201
    owner_articles = upload_owner_resp.json()

    published_resp = test_client.patch(
        f"/api/article-distribution/articles/{owner_articles[0]['id']}/status",
        headers=_headers(owner),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/published",
        },
    )
    invalid_resp = test_client.patch(
        f"/api/article-distribution/articles/{owner_articles[1]['id']}/status",
        headers=_headers(owner),
        json={"publish_status": "invalid"},
    )
    assert published_resp.status_code == 200
    assert invalid_resp.status_code == 200

    page_resp = test_client.get(
        "/api/article-distribution/articles/page",
        headers=_headers(owner),
        params={"page": 2, "page_size": 5, "publication_type": "article"},
    )
    assert page_resp.status_code == 200, page_resp.text
    page_data = page_resp.json()
    assert page_data["total"] == 12
    assert page_data["page"] == 2
    assert page_data["page_size"] == 5
    assert len(page_data["items"]) == 5
    assert [item["title"] for item in page_data["items"]] == [
        "Owner Article 7",
        "Owner Article 6",
        "Owner Article 5",
        "Owner Article 4",
        "Owner Article 3",
    ]
    assert page_data["status_counts"] == {
        "unpublished": 10,
        "published": 1,
        "invalid": 1,
    }

    invalid_page_resp = test_client.get(
        "/api/article-distribution/articles/page",
        headers=_headers(owner),
        params={"publish_status": "invalid", "page": 1, "page_size": 10},
    )
    assert invalid_page_resp.status_code == 200
    invalid_page = invalid_page_resp.json()
    assert invalid_page["total"] == 1
    assert invalid_page["items"][0]["publish_status"] == "invalid"
    assert invalid_page["status_counts"] == {
        "unpublished": 0,
        "published": 0,
        "invalid": 1,
    }

    other_page_resp = test_client.get(
        "/api/article-distribution/articles/page",
        headers=_headers(other),
        params={"page": 1, "page_size": 10},
    )
    assert other_page_resp.status_code == 200
    assert other_page_resp.json()["total"] == 1
    assert other_page_resp.json()["items"][0]["title"] == "Other Article"

    admin_page_resp = test_client.get(
        "/api/article-distribution/articles/page",
        headers=_headers(admin),
        params={"page": 1, "page_size": 20},
    )
    assert admin_page_resp.status_code == 200
    assert admin_page_resp.json()["total"] == 13


def test_unpublished_report_scope_can_be_assigned_to_regular_user(
    test_client, test_db_session: Session
):
    owner_a = _create_user(test_db_session, username="pending_owner_a", name="Owner A")
    owner_b = _create_user(test_db_session, username="pending_owner_b", name="Owner B")
    viewer = _create_user(test_db_session, username="pending_viewer")
    admin = _create_user(test_db_session, username="pending_admin", role=UserRole.ADMIN)

    account_ids: list[int] = []
    for owner, account_name in [(owner_a, "公众号"), (owner_b, "知乎号")]:
        account_resp = test_client.post(
            "/api/article-distribution/accounts",
            headers=_headers(admin),
            json={
                "user_id": owner.id,
                "account_name": account_name,
                "platform": "wechat" if owner.id == owner_a.id else "zhihu",
                "publication_type": "article",
            },
        )
        assert account_resp.status_code == 201
        account_ids.append(account_resp.json()["id"])

    upload_a = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_ids[0],
            "articles": [
                {
                    "title": "A unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-20",
                "project_id": 1,
                },
                {
                    "title": "A published",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-21",
                "project_id": 1,
                },
            ],
        },
    )
    assert upload_a.status_code == 201
    published_article_id = upload_a.json()[1]["id"]
    mark_published = test_client.patch(
        f"/api/article-distribution/articles/{published_article_id}/status",
        headers=_headers(owner_a),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/a-published",
        },
    )
    assert mark_published.status_code == 200

    upload_b = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": account_ids[1],
            "articles": [
                {
                    "title": "B unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-22",
                "project_id": 1,
                }
            ],
        },
    )
    assert upload_b.status_code == 201

    denied_resp = test_client.get(
        "/api/article-distribution/reports/unpublished",
        headers=_headers(viewer),
    )
    assert denied_resp.status_code == 403

    viewer.scope_overrides = auth_service.serialize_scopes(
        [
            *auth_service.get_role_scopes(UserRole.USER),
            auth_service.SCOPE_ARTICLE_DISTRIBUTION_REPORT_READ,
        ]
    )
    test_db_session.commit()
    test_db_session.refresh(viewer)

    report_resp = test_client.get(
        "/api/article-distribution/reports/unpublished",
        headers=_headers(viewer),
    )

    assert report_resp.status_code == 200, report_resp.text
    report = report_resp.json()
    assert report["summary"] == {
        "total_users": 2,
        "unpublished_users": 2,
        "published_articles": 1,
        "unpublished_articles": 2,
        "invalid_articles": 0,
        "inactive_account_articles": 0,
        "read_count": 0,
        "like_count": 0,
        "favorite_count": 0,
        "share_count": 0,
    }
    data = report["users"]
    assert [item["user_id"] for item in data] == [owner_a.id, owner_b.id]
    assert [item["remaining_count"] for item in data] == [1, 1]
    assert data[0]["published_count"] == 1
    assert data[0]["platform_summaries"] == []
    assert data[0]["articles"] == []

    detail_resp = test_client.get(
        f"/api/article-distribution/reports/unpublished/users/{owner_a.id}",
        headers=_headers(viewer),
    )
    assert detail_resp.status_code == 200
    detail = detail_resp.json()
    assert detail["platform_summaries"][0]["published_count"] == 1
    assert detail["platform_summaries"][0]["latest_published_url"] == "https://example.com/a-published"
    assert detail["articles"][0]["title"] == "A unpublished"
    assert detail["articles"][0]["markdown_content"] == "body"
    assert detail["articles"][0]["account_name"] == "公众号"
    assert detail["articles"][0]["platform"] == "wechat"
    assert detail["articles"][0]["publish_status"] == "unpublished"
    assert detail["articles"][1]["publish_status"] == "published"
    assert "source" not in detail["articles"][0]

    detail_b_resp = test_client.get(
        f"/api/article-distribution/reports/unpublished/users/{owner_b.id}",
        headers=_headers(viewer),
    )
    assert detail_b_resp.status_code == 200
    assert detail_b_resp.json()["articles"][0]["title"] == "B unpublished"

    dao = ProjectManagementDAO(test_db_session)
    default_project = dao.get_project_by_name("AIFC")
    default_theme = dao.get_theme_by_name("AI")
    assert default_project is not None
    assert default_theme is not None
    second_project = dao.create_project(
        Project(name="Second Project", code="BBBBBBBB", is_active=True)
    )
    dao.replace_project_themes(second_project.id, [default_theme.id])
    dao.add_user_project(owner_a.id, second_project.id)
    second_project_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner_a.id,
            "account_name": "第二项目公众号",
            "platform": "wechat",
            "publication_type": "article",
            "project_id": second_project.id,
            "theme_id": default_theme.id,
        },
    )
    assert second_project_account_resp.status_code == 201

    second_project_upload = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": second_project_account_resp.json()["id"],
            "articles": [
                {
                    "title": "Second project published",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-23",
                    "project_id": second_project.id,
                },
            ],
        },
    )
    assert second_project_upload.status_code == 201
    second_project_article_id = second_project_upload.json()[0]["id"]
    second_project_mark_published = test_client.patch(
        f"/api/article-distribution/articles/{second_project_article_id}/status",
        headers=_headers(owner_a),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/second-project",
        },
    )
    assert second_project_mark_published.status_code == 200

    public_resp = test_client.get(
        f"/api/article-distribution/public/dashboard/{default_project.code}"
    )
    assert public_resp.status_code == 200, public_resp.text
    public_report = public_resp.json()
    assert public_report["summary"] == {
        "total_users": 2,
        "unpublished_users": 2,
        "published_articles": 1,
        "unpublished_articles": 2,
        "invalid_articles": 0,
        "inactive_account_articles": 0,
        "read_count": 0,
        "like_count": 0,
        "favorite_count": 0,
        "share_count": 0,
    }
    assert public_report["total"] == 1
    assert public_report["page"] == 1
    assert public_report["page_size"] == 10
    assert public_report["articles"] == [
        {
            "id": published_article_id,
            "title": "A published",
            "published_at": "2026-05-21",
            "published_url": "https://example.com/a-published",
            "account_name": "公众号",
            "platform": "wechat",
            "publication_type": "article",
            "latest_traffic_stat": None,
        }
    ]

    second_project_public_resp = test_client.get(
        "/api/article-distribution/public/dashboard/BBBBBBBB"
    )
    assert second_project_public_resp.status_code == 200
    second_project_public_report = second_project_public_resp.json()
    assert second_project_public_report["summary"]["published_articles"] == 1
    assert second_project_public_report["summary"]["unpublished_articles"] == 0
    assert second_project_public_report["total"] == 1
    assert (
        second_project_public_report["articles"][0]["title"]
        == "Second project published"
    )

    public_filtered_resp = test_client.get(
        f"/api/article-distribution/public/dashboard/{default_project.code}",
        params={
            "publication_type": "video",
            "scheduled_from": "2026-05-20",
            "scheduled_to": "2026-05-22",
        },
    )
    assert public_filtered_resp.status_code == 200
    assert public_filtered_resp.json()["summary"]["total_users"] == 0
    assert public_filtered_resp.json()["total"] == 0
    assert public_filtered_resp.json()["articles"] == []


def test_unpublished_report_tracks_inactive_account_articles_separately(
    test_client, test_db_session: Session
):
    owner = _create_user(test_db_session, username="inactive_report_owner")
    admin = _create_user(
        test_db_session, username="inactive_report_admin", role=UserRole.ADMIN
    )

    active_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "启用号",
            "platform": "wechat",
            "publication_type": "article",
        },
    )
    inactive_account_resp = test_client.post(
        "/api/article-distribution/accounts",
        headers=_headers(admin),
        json={
            "user_id": owner.id,
            "account_name": "停用号",
            "platform": "wechat",
            "publication_type": "image_text",
        },
    )
    assert active_account_resp.status_code == 201
    assert inactive_account_resp.status_code == 201
    active_account_id = active_account_resp.json()["id"]
    inactive_account_id = inactive_account_resp.json()["id"]

    active_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": active_account_id,
            "articles": [
                {
                    "title": "Active unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-26",
                "project_id": 1,
                }
            ],
        },
    )
    inactive_upload_resp = test_client.post(
        "/api/admin/article-distribution/articles",
        headers=_headers(admin),
        json={
            "account_id": inactive_account_id,
            "articles": [
                {
                    "title": "Inactive unpublished",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-26",
                "project_id": 1,
                },
                {
                    "title": "Inactive published",
                    "markdown_content": "body",
                    "scheduled_date": "2026-05-27",
                "project_id": 1,
                },
            ],
        },
    )
    assert active_upload_resp.status_code == 201
    assert inactive_upload_resp.status_code == 201
    inactive_published_id = inactive_upload_resp.json()[1]["id"]

    published_resp = test_client.patch(
        f"/api/article-distribution/articles/{inactive_published_id}/status",
        headers=_headers(owner),
        json={
            "publish_status": "published",
            "published_url": "https://example.com/inactive-published",
        },
    )
    assert published_resp.status_code == 200

    deactivate_resp = test_client.patch(
        f"/api/article-distribution/accounts/{inactive_account_id}",
        headers=_headers(owner),
        json={"is_active": False},
    )
    assert deactivate_resp.status_code == 200

    active_report_resp = test_client.get(
        "/api/article-distribution/reports/unpublished",
        headers=_headers(admin),
    )
    assert active_report_resp.status_code == 200
    active_report = active_report_resp.json()
    assert active_report["summary"] == {
        "total_users": 1,
        "unpublished_users": 1,
        "published_articles": 0,
        "unpublished_articles": 1,
        "invalid_articles": 0,
        "inactive_account_articles": 0,
        "read_count": 0,
        "like_count": 0,
        "favorite_count": 0,
        "share_count": 0,
    }
    assert active_report["users"][0]["remaining_count"] == 1
    assert active_report["users"][0]["articles"] == []
    active_detail_resp = test_client.get(
        f"/api/article-distribution/reports/unpublished/users/{owner.id}",
        headers=_headers(admin),
    )
    assert active_detail_resp.status_code == 200
    assert [article["account_is_active"] for article in active_detail_resp.json()["articles"]] == [True]

    all_report_resp = test_client.get(
        "/api/article-distribution/reports/unpublished",
        headers=_headers(admin),
        params={"account_status": "all"},
    )
    assert all_report_resp.status_code == 200
    all_report = all_report_resp.json()
    assert all_report["summary"] == {
        "total_users": 1,
        "unpublished_users": 1,
        "published_articles": 1,
        "unpublished_articles": 1,
        "invalid_articles": 0,
        "inactive_account_articles": 2,
        "read_count": 0,
        "like_count": 0,
        "favorite_count": 0,
        "share_count": 0,
    }
    all_user = all_report["users"][0]
    assert all_user["remaining_count"] == 1
    assert all_user["platform_summaries"] == []
    assert all_user["articles"] == []
    all_detail_resp = test_client.get(
        f"/api/article-distribution/reports/unpublished/users/{owner.id}",
        headers=_headers(admin),
        params={"account_status": "all"},
    )
    assert all_detail_resp.status_code == 200
    all_detail = all_detail_resp.json()
    inactive_summary = next(
        summary
        for summary in all_detail["platform_summaries"]
        if summary["account_id"] == inactive_account_id
    )
    assert inactive_summary["account_is_active"] is False
    assert inactive_summary["unpublished_count"] == 0
    assert [article["account_is_active"] for article in all_detail["articles"]] == [
        True,
        False,
        False,
    ]

    inactive_report_resp = test_client.get(
        "/api/article-distribution/reports/unpublished",
        headers=_headers(admin),
        params={"account_status": "inactive"},
    )
    assert inactive_report_resp.status_code == 200
    inactive_report = inactive_report_resp.json()
    assert inactive_report["summary"] == {
        "total_users": 1,
        "unpublished_users": 0,
        "published_articles": 1,
        "unpublished_articles": 0,
        "invalid_articles": 0,
        "inactive_account_articles": 2,
        "read_count": 0,
        "like_count": 0,
        "favorite_count": 0,
        "share_count": 0,
    }
    assert inactive_report["users"][0]["remaining_count"] == 0
    inactive_detail_resp = test_client.get(
        f"/api/article-distribution/reports/unpublished/users/{owner.id}",
        headers=_headers(admin),
        params={"account_status": "inactive"},
    )
    assert inactive_detail_resp.status_code == 200
    assert inactive_detail_resp.json()["remaining_count"] == 0
